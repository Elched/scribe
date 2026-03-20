"""
import_config_xlsx.py — Initialise SCRIBE depuis le fichier Excel de configuration
Usage : python import_config_xlsx.py [chemin/vers/config.xlsx]

Lit les 5 onglets :
  ETABLISSEMENT    → config.xml + config.js + auth.py
  DIRECTEURS       → config.js (directeurs)
  TELEPHONIE       → config.js (annuaire normal + secours)
  UF_INCIDENTS     → base de données (unites_fonctionnelles)
  SERVICES_CAPACITE→ base de données (capacite_referentiel)
"""
import sys, os, re, json, hashlib, xml.etree.ElementTree as ET
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))

try:
    import openpyxl
except ImportError:
    print("[!] openpyxl manquant. Lancez : pip install openpyxl")
    sys.exit(1)

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "SCRIBE_config_etablissement.xlsx"

if not os.path.exists(XLSX_PATH):
    print(f"[!] Fichier introuvable : {XLSX_PATH}")
    print("    Placez SCRIBE_config_etablissement.xlsx dans ce dossier ou passez le chemin en argument.")
    sys.exit(1)

print(f"""
╔══════════════════════════════════════════════════════════════╗
║  SCRIBE — Import configuration depuis Excel                  ║
╚══════════════════════════════════════════════════════════════╝
  Fichier : {XLSX_PATH}
""")

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ── Lire ETABLISSEMENT ──────────────────────────────────────────────────────
ws_etab = wb["ETABLISSEMENT"]
params = {}
for row in ws_etab.iter_rows(min_row=4, values_only=True):
    if row[0] and row[1] is not None:
        params[str(row[0]).strip()] = str(row[1]).strip() if row[1] else ""

# Lire les sites (après les lignes de paramètres)
sites = []
in_sites = False
for row in ws_etab.iter_rows(min_row=4, values_only=True):
    if row[0] and "site" in str(row[0]).lower() and "principal" in str(row[0]).lower().replace("é","e"):
        in_sites = True
    if in_sites and row[0] and row[2]:  # a un nom et une latitude
        try:
            float(str(row[2]))
            sites.append({
                "nom": str(row[0]).strip(),
                "adresse": str(row[1] or "").strip(),
                "latitude": str(row[2]).strip(),
                "longitude": str(row[3] or "").strip(),
                "telephone_garde": str(row[4] or "").strip(),
            })
        except (ValueError, TypeError):
            pass

# Parser les sites depuis les lignes après le titre "Sites géographiques"
sites = []
site_section = False
for row in ws_etab.iter_rows(min_row=1, values_only=True):
    if row[0] and "Sites" in str(row[0]):
        site_section = True
        continue
    if site_section and row[0] and row[0] not in ["Nom du site"] and str(row[0]).strip():
        try:
            if row[2]:
                float(str(row[2]))
                sites.append({
                    "nom": str(row[0]).strip(),
                    "adresse": str(row[1] or "").strip(),
                    "latitude": str(row[2]).strip(),
                    "longitude": str(row[3] or "").strip(),
                    "telephone_garde": str(row[4] or "").strip(),
                })
        except (ValueError, TypeError):
            pass

nom_etab = params.get("NOM_ETABLISSEMENT", "Mon Établissement")
sigle    = params.get("SIGLE", "ETB")
finess   = params.get("FINESS", "")
langue   = params.get("LANGUE", "fr")
login    = params.get("LOGIN_ADMIN", "dircrise")
password = params.get("MOT_DE_PASSE", "Scribe2026!")
nom_admin= params.get("NOM_AFFICHE_ADMIN", "Directeur de Crise")
ia_fourn = params.get("FOURNISSEUR_IA", "albert")
ia_key   = params.get("CLE_API_IA", "")
ia_model = params.get("MODELE_IA", "")
ia_url   = params.get("URL_BASE_IA", "")
fed_on   = params.get("FEDERATION_ACTIVE", "false")
col_url  = params.get("COLLECTEUR_URL", "")
col_tok  = params.get("COLLECTEUR_TOKEN", "")
sync_cr  = params.get("SYNC_CRISE", "false")
sync_sa  = params.get("SYNC_SANITAIRE", "false")

print(f"  Établissement : {nom_etab} ({sigle})")
print(f"  Sites trouvés : {len(sites)}")
print(f"  Admin         : {login}")
print(f"  IA            : {ia_fourn}")

# ── Lire DIRECTEURS ─────────────────────────────────────────────────────────
ws_dirs = wb["DIRECTEURS"]
directeurs = []
for row in ws_dirs.iter_rows(min_row=4, values_only=True):
    if row[0] and str(row[0]).strip() and str(row[0]).strip() not in ["Nom Prénom"]:
        directeurs.append({
            "nom": str(row[0]).strip(),
            "fonction": str(row[1] or "").strip(),
            "abreviation": str(row[2] or "").strip(),
        })
print(f"  Directeurs    : {len(directeurs)}")

# ── Lire TELEPHONIE ─────────────────────────────────────────────────────────
ws_tel = wb["TELEPHONIE"]
annuaire_normal  = []
annuaire_secours = []
current = None
for row in ws_tel.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    cell = str(row[0]).strip()
    if "CONTACTS NOMINAUX" in cell.upper():
        current = "normal"
        continue
    if "CONTACTS SECOURS" in cell.upper():
        current = "secours"
        continue
    if cell in ["Service","N° Interne / IP"]:
        continue
    if current == "normal" and row[1]:
        annuaire_normal.append({
            "service": cell,
            "local": str(row[1] or "").strip(),
            "tel": str(row[2] or "").strip(),
        })
    elif current == "secours" and row[1]:
        annuaire_secours.append({
            "service": cell,
            "local": str(row[1] or "").strip(),
            "tel": str(row[2] or "").strip(),
            "note": str(row[5] or "").strip(),
        })
print(f"  Contacts norm : {len(annuaire_normal)}")
print(f"  Contacts sec. : {len(annuaire_secours)}")

# ── Générer config.xml ──────────────────────────────────────────────────────
def esc(s):
    return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

sites_xml = ""
for s in sites:
    sites_xml += f"""    <site>
      <nom>{esc(s['nom'])}</nom>
      <adresse>{esc(s['adresse'])}</adresse>
      <latitude>{esc(s['latitude'])}</latitude>
      <longitude>{esc(s['longitude'])}</longitude>
      <telephone_garde>{esc(s['telephone_garde'])}</telephone_garde>
    </site>
"""

dirs_xml = ""
for d in directeurs:
    dirs_xml += f"""    <directeur>
      <nom>{esc(d['nom'])}</nom>
      <fonction>{esc(d['fonction'])}</fonction>
      <abreviation>{esc(d['abreviation'])}</abreviation>
    </directeur>
"""

xml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<!-- Généré par import_config_xlsx.py le {datetime.now().strftime('%d/%m/%Y %H:%M')} -->
<scribe>
  <etablissement>
    <nom>{esc(nom_etab)}</nom>
    <sigle>{esc(sigle)}</sigle>
    <finess>{esc(finess)}</finess>
  </etablissement>
  <admin>
    <login>{esc(login)}</login>
    <password>{esc(password)}</password>
    <nom_affiche>{esc(nom_admin)}</nom_affiche>
  </admin>
  <sites>
{sites_xml}  </sites>
  <directeurs>
{dirs_xml}  </directeurs>
  <langue>{esc(langue)}</langue>
  <ia>
    <fournisseur>{esc(ia_fourn)}</fournisseur>
    <cle_api>{esc(ia_key)}</cle_api>
    <modele>{esc(ia_model)}</modele>
    <url_base>{esc(ia_url)}</url_base>
  </ia>
  <federation>
    <enabled>{esc(fed_on)}</enabled>
    <collecteur_url>{esc(col_url)}</collecteur_url>
    <token>{esc(col_tok)}</token>
    <intervalle_secondes>30</intervalle_secondes>
    <share_details>true</share_details>
    <share_min_urgency>1</share_min_urgency>
    <sync_crise>{esc(sync_cr)}</sync_crise>
    <sync_sanitaire>{esc(sync_sa)}</sync_sanitaire>
    <share_capacite_details>true</share_capacite_details>
  </federation>
</scribe>
"""

xml_path = os.path.join(os.path.dirname(__file__), "config.xml")
with open(xml_path, "w", encoding="utf-8") as f:
    f.write(xml_content)
print(f"\n  [✓] config.xml généré")

# ── Lancer setup.py ──────────────────────────────────────────────────────────
import subprocess
r = subprocess.run([sys.executable, "setup.py", "config.xml"],
                  capture_output=False)
if r.returncode != 0:
    print("  [!] setup.py a échoué")
    sys.exit(1)

# ── Importer les UF ─────────────────────────────────────────────────────────
from app.database import SessionLocal, Base, engine
from app.models import Hospital, UniteFonctionnelle, CapaciteReferentiel
Base.metadata.create_all(bind=engine)
db = SessionLocal()

site_map = {h.nom: h.id for h in db.query(Hospital).all()}

ws_uf = wb["UF_INCIDENTS"]
uf_count = 0
for row in ws_uf.iter_rows(min_row=4, values_only=True):
    if not row[0] or not row[1] or str(row[0]).strip() == "Code UF":
        continue
    if row[4] and str(row[4]).strip().upper() == "N":
        continue
    code  = str(row[0]).strip()
    lib   = str(row[1]).strip()
    pole  = str(row[2] or "").strip()
    site_nom = str(row[3] or "").strip()
    hospital_id = site_map.get(site_nom)
    if not hospital_id:
        # Cherche le premier site par défaut
        hospital_id = list(site_map.values())[0] if site_map else 1
    exists = db.query(UniteFonctionnelle).filter_by(code_uf=code, hospital_id=hospital_id).first()
    if not exists:
        db.add(UniteFonctionnelle(
            code_uf=code, libelle=lib, pole=pole, hospital_id=hospital_id
        ))
        uf_count += 1

db.commit()
print(f"  [✓] {uf_count} UF importées (onglet UF_INCIDENTS)")

# ── Importer le référentiel capacitaire ─────────────────────────────────────
ws_cap = wb["SERVICES_CAPACITE"]
cap_count = 0
for row in ws_cap.iter_rows(min_row=4, values_only=True):
    if not row[0] or str(row[0]).strip() == "Service":
        continue
    nom   = str(row[0]).strip()
    code  = str(row[1] or "").strip() or None
    pole  = str(row[2] or "").strip()
    site  = str(row[3] or "").strip()
    capa  = int(row[4] or 0)
    t1    = int(row[5] or 0)
    t2    = int(row[6] or 0)
    h     = str(row[7] or "O").strip().upper() == "O"
    f     = str(row[8] or "O").strip().upper() == "O"
    ind   = str(row[9] or "O").strip().upper() == "O"
    tel   = str(row[10] or "").strip()
    ordre = int(row[11] or 99)
    exists = db.query(CapaciteReferentiel).filter_by(
        service_nom=nom, site=site
    ).first()
    if not exists:
        db.add(CapaciteReferentiel(
            service_nom=nom, uf_code=code, pole=pole, site=site,
            capacite_totale=capa, tension_1=t1, tension_2=t2,
            accept_homme=h, accept_femme=f, accept_indiffer=ind,
            telephone_cadre=tel, ordre_affichage=ordre,
        ))
        cap_count += 1

db.commit()
total_cap = db.query(CapaciteReferentiel).count()
db.close()
print(f"  [✓] {cap_count} services capacitaires importés (onglet SERVICES_CAPACITE)")
print(f"  [✓] Total référentiel capacitaire : {total_cap} services")

print(f"""
╔══════════════════════════════════════════════════════════════╗
║  Import terminé !                                            ║
╚══════════════════════════════════════════════════════════════╝

  Établissement  : {nom_etab} ({sigle})
  Sites          : {len(sites)}
  Directeurs     : {len(directeurs)}
  UF incidents   : {uf_count} importées
  Capacité       : {cap_count} services importés

  ─── DÉMARRAGE ─────────────────────────────────────────────
  $ python main.py
  Puis ouvrez : http://localhost:8000
  Login : {login}  /  Mot de passe : (voir onglet ETABLISSEMENT)
""")
